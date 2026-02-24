import json
from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime, timedelta
import os

# --- Firestore Setup Imports ---
import firebase_admin
from firebase_admin import credentials, firestore, storage
from firebase_admin import initialize_app, firestore, credentials, auth
from werkzeug.utils import secure_filename
from flask import send_file
from flask import Flask, render_template, request, redirect, url_for, flash, after_this_request

import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Define the global variables provided by the environment
__app_id = os.environ.get('APP_ID', 'default-app-id')
__firebase_config = os.environ.get('FIREBASE_CONFIG', '{}')
__initial_auth_token = os.environ.get('AUTH_TOKEN', None)

# --- Flask App Initialization and Firestore/Auth Setup ---

app = Flask(__name__)
app.secret_key = 'gulmohour-secret'

EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'sender_email': 'contact.gulmohurfashion@gmail.com',  # Your Gmail address
    'sender_password': 'jydv jtez uhce lizm',        # Your App Password (16 chars)
    'sender_name': 'Gulmohur Fashion'
}

WHATSAPP_CONFIG = {
    'business_number': '919810137621',  # Format: CountryCode + Number (no + or spaces)
}

# --- Firestore Setup Imports ---
import firebase_admin
from firebase_admin import credentials, firestore, storage
import json
import os

# Global variables
db = None
bucket = None

def init_firebase():
    """Initialize Firebase with proper error handling"""
    global db, bucket
    
    try:
        # Check if app already exists
        try:
            app = firebase_admin.get_app()
            print("ℹ️ Firebase app already initialized")
        except ValueError:
            # App doesn't exist, initialize it
            firebase_cred_json = os.environ.get('FIREBASE_CREDENTIALS')
            
            if firebase_cred_json:
                print("🔄 Loading Firebase from environment variable...")
                print(f"📏 Credential length: {len(firebase_cred_json)}")
                
                # Parse JSON
                cred_dict = json.loads(firebase_cred_json)
                print(f"✅ JSON parsed successfully. Project: {cred_dict.get('project_id')}")
                
                cred = credentials.Certificate(cred_dict)
            else:
                print("🔄 Loading Firebase from local file...")
                # Path to your firebase-key.json (same folder as app.py)
                cred_path = os.path.join(os.path.dirname(__file__), "firebase-key.json")
                
                # Load credentials from the JSON file
                cred = credentials.Certificate(cred_path)
            
            firebase_admin.initialize_app(cred, {
                'storageBucket': 'gulmohur-fashion.firebasestorage.app'
            })
            print("✅ Firebase app initialized")
        
        # Create clients
        db = firestore.client()
        print(f"✅ Firestore client created: {db is not None}")

        # Initialize Firebase Storage bucket
        try:
            bucket = storage.bucket()
            print("Firebase Storage initialized successfully")
        except Exception as e:
            print(f"Firebase Storage initialization failed: {e}")
            bucket = None
        
        return True
        
    except json.JSONDecodeError as e:
        print(f"❌ JSON parsing error: {e}")
        print(f"🔍 First 100 chars of credential: {os.environ.get('FIREBASE_CREDENTIALS', '')[:100]}")
        db = None
        bucket = None
        return False
        
    except Exception as e:
        print(f"❌ Firebase initialization failed: {e}")
        print(f"🔍 Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        db = None
        bucket = None
        return False

# Initialize Firebase when module loads
print("=" * 50)
print("🚀 Starting Firebase initialization...")
firebase_initialized = init_firebase()
print(f"📊 Firebase initialized: {firebase_initialized}")
print(f"📊 db is None: {db is None}")
print(f"📊 bucket is None: {bucket is None}")
print("=" * 50)

# Global variable to store current user's ID
CURRENT_USER_ID = None

def initialize_auth():
    """Initializes authentication and sets the CURRENT_USER_ID."""
    global CURRENT_USER_ID, db
    
    if db is None:
        print("⚠️ WARNING: db is None in initialize_auth!")
        CURRENT_USER_ID = "no_db_connection"
    else:
        try:
            if __initial_auth_token:
                CURRENT_USER_ID = __app_id
            else:
                CURRENT_USER_ID = "anonymous_user_default_id"
        except Exception as e:
            print(f"Authentication setup failed: {e}")
            CURRENT_USER_ID = "anonymous_user_failed_auth"
    
    print(f"Using CURRENT_USER_ID: {CURRENT_USER_ID}")

# --- Helper Functions ---

# --- Firestore Helper Functions ---



def get_collection_ref(collection_name):
    """Returns the Firestore CollectionReference for the current user's private data."""
    global db
    
    if db is None:
        print(f"❌ ERROR: db is None when accessing collection '{collection_name}'")
        print("🔄 Attempting to reinitialize Firebase...")
        init_firebase()
        
        if db is None:
            print(f"❌ CRITICAL: Still cannot access database after reinit!")
            raise Exception("Firebase database not initialized. Check FIREBASE_CREDENTIALS environment variable.")
    
    return db.collection(collection_name)

# # 1. Initialize Firebase App
# try:
#     # load_dotenv()

#     firebase_key_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")

#     cred = credentials.Certificate(firebase_key_path)

#     firebase_admin.initialize_app(cred, {
#     'storageBucket': 'gulmohour-boutique.firebasestorage.app'
# })
#     db = firestore.client()
# except Exception as e:
#     print(f"Failed to initialize Firebase: {e}")
#     db = None

# # Global variable to store current user's ID
# CURRENT_USER_ID = None

# def initialize_auth():
#     """Initializes authentication and sets the CURRENT_USER_ID."""
#     global CURRENT_USER_ID
#     if db:
#         try:
#             if __initial_auth_token:
#                 CURRENT_USER_ID = __app_id
#             else:
#                 CURRENT_USER_ID = "anonymous_user_default_id"
#         except Exception as e:
#             print(f"Authentication setup failed: {e}")
#             CURRENT_USER_ID = "anonymous_user_failed_auth"
#     else:
#         CURRENT_USER_ID = "no_db_connection"
#     print(f"Using CURRENT_USER_ID: {CURRENT_USER_ID}")

# Image upload configuration
UPLOAD_FOLDER = 'uploads/order_images'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB


# Shop Details - Add this constant near the top of your file after the app initialization
SHOP_DETAILS = {
    'owner_name': 'Parul Gupta',
    'shop_name': 'Gulmohur',
    'tagline': 'Fabric for all Seasons',
    'phone1': '9810137621',
    'phone2': '9818326194',
    'address': 'R3134, Second Floor (Level 1), 65th Avenue, SPR opp DPS International Gurgaon -122102'
}


os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# -----------------------------
# Helper: Safe payments parsing
# -----------------------------
def parse_payments(payments):
    """Ensure payments is always a list of dicts, even if stored as a JSON string."""
    if isinstance(payments, str):
        try:
            payments = json.loads(payments)
        except:
            payments = []
    if not isinstance(payments, list):
        payments = []
    return [p for p in payments if isinstance(p, dict)]


def get_customer_by_id(customer_id):
    """Fetches a single customer document by ID."""
    customers_ref = get_collection_ref('customers')
    if not customers_ref: return None
    
    doc = customers_ref.document(customer_id).get()
    if doc.exists:
        customer_data = doc.to_dict()
        customer_data['id'] = doc.id
        return customer_data
    return None

def get_order_by_id(order_id):
    """Fetches a single order document by ID."""
    orders_ref = get_collection_ref('orders')
    if not orders_ref: return None
    
    doc = orders_ref.document(order_id).get()
    if doc.exists:
        order_data = doc.to_dict()
        order_data['id'] = doc.id
        return order_data
    return None

# ========================================
# SECTION 3: ADD IMAGE HELPER FUNCTIONS (before routes)
# ========================================

def upload_image_to_firebase(file, order_id):
    """Upload image to Firebase Storage and return public URL"""
    if not bucket:
        return None
    
    try:
        filename = secure_filename(file.filename)
        blob_name = f"orders/{order_id}/{datetime.now().timestamp()}_{filename}"
        blob = bucket.blob(blob_name)
        
        blob.upload_from_file(file, content_type=file.content_type)
        blob.make_public()
        
        return blob.public_url
    except Exception as e:
        print(f"Error uploading to Firebase Storage: {e}")
        return None

def generate_receipt_pdf(order, customer):
    """Generate professional receipt PDF matching exact format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=25, leftMargin=25, 
                           topMargin=20, bottomMargin=20)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=9,
        leading=11
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_CENTER,
        spaceAfter=2
    )
    
    address_style = ParagraphStyle(
        'Address',
        parent=styles['Normal'],
        fontSize=7,
        alignment=TA_CENTER,
        spaceAfter=8
    )
    
    # Header with owner name, bill receipt, and phones
    header_data = [[
        Paragraph(SHOP_DETAILS['owner_name'], header_style),
        Paragraph('<b>Bill Receipt Book</b>', header_style),
        Paragraph(f"{SHOP_DETAILS['phone1']}<br/>{SHOP_DETAILS['phone2']}", header_style)
    ]]
    
    header_table = Table(header_data, colWidths=[1.8*inch, 2.2*inch, 1.8*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 5))
    
    # ✅ LARGE CENTERED LOGO
    try:
        from reportlab.platypus import Image
        #logo_path = 'static/gulmohur_logo.jpg'
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'gulmohur_logo.jpg')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.5*inch, height=1.5*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 5))  # Reduced space after logo
    except Exception as e:
        print(f"Logo not found: {e}")
    
    # Tagline and Address - TIGHT SPACING
    elements.append(Paragraph(SHOP_DETAILS['tagline'], subtitle_style))
    elements.append(Spacer(1, 2))  # Minimal space
    elements.append(Paragraph(SHOP_DETAILS['address'], address_style))
    elements.append(Spacer(1, 8))
    
    # ✅ Customer details box WITH DOTTED LINES
    customer_address = ""
    if customer.get('address'):
        addr = customer['address']
        parts = [addr.get('house', ''), addr.get('locality', ''), 
                addr.get('city', ''), addr.get('state', ''), addr.get('pin', '')]
        customer_address = ', '.join([p for p in parts if p])
    
    # Create dotted line style
    dotted_style = ParagraphStyle(
        'Dotted',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    
    customer_info_data = [
        [Paragraph('Customer Name & Address:', dotted_style), Paragraph(f"<b>S. No. #{order.get('order_number', 'N/A')}</b>", dotted_style)],
        [Paragraph(f"<b>{customer.get('name', 'N/A')}</b>", dotted_style), ''],
        [Paragraph('.' * 80, dotted_style), ''],
        [Paragraph(customer_address[:60] if customer_address else '.' * 80, dotted_style), ''],
        [Paragraph('.' * 80, dotted_style), ''],
        [Paragraph(f"Cell/Contact No: {customer.get('phone', 'N/A')}", dotted_style), Paragraph(f"Booking Date: {datetime.now().strftime('%d-%m-%Y')}", dotted_style)],
        [Paragraph('.' * 80, dotted_style), Paragraph(f"Delivery Date: {order.get('delivery_date', 'N/A')}", dotted_style)]
    ]
    
    customer_table = Table(customer_info_data, colWidths=[3.5*inch, 2.3*inch])
    customer_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 8))
    
    # ✅ MEASUREMENTS TABLE - FIXED FORMAT (Particulars | Design | Comments)
    # ✅ MEASUREMENTS TABLE - NO GRID LINES in Design/Comments columns
    measurements = order.get('measurements', {})
    notes = order.get('notes', '')
    
    measurements_data = [['Particulars', 'Design', 'Comments/Special Notes']]
    
    # Measurement fields
    measure_fields = [
        ('Length:', 'length'), ('Shoulder:', 'shoulder'),
        ('Armhole:', 'armhole'), ('Upper Chest:', 'upper_chest'),
        ('Chest:', 'chest'), ('Waist:', 'waist'),
        ('Stomach:', 'stomach'), ('Hips:', 'hips'),
        ('Front/Back Cross:', 'front_back_cross'),
        ('Sleeve Length:', 'sleeve_length'),
        ('Neck (F&B):', 'neck_front_back'),
        ('Dart Point:', 'dart_point'),
    ]
    
    # Add measurements - show value with label in Particulars
    for idx, (label, key) in enumerate(measure_fields):
        value = measurements.get(key, '')
        particulars_text = f"{label} {value}" if value else label
        if idx == 0 and notes:
            measurements_data.append([particulars_text, '', notes[:100]])
        else:
            measurements_data.append([particulars_text, '', ''])
    
    # Pants section
    pants = measurements.get('pants', {})
    measurements_data.append(['Pants', '', ''])
    pants_fields = [('Length', 'length'), ('Waist', 'waist'), ('Thigh', 'thigh'), ('Mori', 'mori'), ('Calf', 'calf')]
    for label, key in pants_fields:
        value = pants.get(key, '')
        measurements_data.append([f"  - {label}: {value}" if value else f"  - {label}:", '', ''])
    
    # ✅ REDUCED Particulars column width, NO GRID on Design/Comments
    measurements_table = Table(measurements_data, colWidths=[1.8*inch, 1.2*inch, 2.8*inch])
    measurements_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        # ✅ Only vertical grid lines between columns
        ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.grey),  # Line after Particulars
        ('LINEAFTER', (1, 0), (1, -1), 0.5, colors.grey),  # Line after Design
        # ✅ Horizontal line only for header
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e0e0')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(measurements_table)
    elements.append(Spacer(1, 8))
    
    # Financial summary table
    financial_data = [
        ['Total', f"₹{order.get('price', 0):.0f}"],
        ['Advance received', f"₹{order.get('advance_received', 0):.0f}"],
        ['Balance Amount', f"₹{order.get('balance_amount', 0):.0f}"]
    ]
    
    financial_table = Table(financial_data, colWidths=[4.3*inch, 1.5*inch])
    financial_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
    ]))
    elements.append(financial_table)
    elements.append(Spacer(1, 8))
    
    # ALL DISCLAIMER POINTS
    terms_style = ParagraphStyle('Terms', parent=styles['Normal'], fontSize=6, leading=8)
    
    disclaimer_points = [
        "Stitching orders are as per measurements provided by the customer",
        "*Customers are responsible for providing accurate measurements.",
        "*50% advance at the time of placement of stitching orders",
        "*Delivery timeline will be provided at the time of order confirmation",
        "*Minor alterations due to fitting issues will be provided free of charge within a specified time frame after delivery.",
        "Any major alterations/changes in design may incur additional charges",
        "*Shipping charges if applicable will be communicated upfront",
        "*Design, patterns & other intellectual property used in stitching orders remain the property of Gulmohur"
    ]
    
    for point in disclaimer_points:
        elements.append(Paragraph(point, terms_style))
    
    elements.append(Spacer(1, 8))
    
    # UPI QR Code
    try:
        from reportlab.platypus import Image as RLImage
        import qrcode
        from io import BytesIO
        
        upi_string = "upi://pay?pa=9711217193@ptsbi&pn=Nitin Gupta&cu=INR"
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(upi_string)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        qr_image = RLImage(qr_buffer, width=1.5*inch, height=1.5*inch)
        
        upi_text = Paragraph(
            '<b>paytm♥UPI♥</b><br/><br/>'
            '<b>Scan to Pay via UPI</b><br/>'
            'paytm/googlePay/PhonePe<br/><br/>'
            'Nitin Gupta<br/>'
            '#9711217193@ptsbi', 
            ParagraphStyle('UPI', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER)
        )
        
        upi_data = [[qr_image, upi_text]]
        upi_table = Table(upi_data, colWidths=[2*inch, 3.8*inch])
        upi_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(upi_table)
    except Exception as e:
        print(f"Error adding UPI QR: {e}")
        elements.append(Paragraph(
            '<b>UPI Payment:</b> 9711217193@ptsbi (Nitin Gupta)', 
            ParagraphStyle('UPI', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)
        ))
        
    doc.build(elements)
    buffer.seek(0)
    return buffer

@app.route('/order/email-receipt/<order_id>', methods=['POST'])
def email_receipt(order_id):
    """Send receipt via email - same PDF as download"""
    try:
        from flask import jsonify
        
        order = get_order_by_id(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'Order not found'}), 404
        
        customer = get_customer_by_id(order['customer_id'])
        if not customer:
            return jsonify({'success': False, 'message': 'Customer not found'}), 404
        
        customer_email = customer.get('email')
        if not customer_email:
            return jsonify({'success': False, 'message': 'Customer email not found. Please add email to customer profile.'}), 400
        
        # Import email libraries
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        # Generate the SAME PDF as download receipt
        buffer = generate_receipt_pdf(order, customer)
        
        # Create email
        msg = MIMEMultipart()
        msg['From'] = f"{EMAIL_CONFIG['sender_name']} <{EMAIL_CONFIG['sender_email']}>"
        msg['To'] = customer_email
        msg['Subject'] = f"Order Receipt - #{order.get('order_number', 'N/A')} - Gulmohur"
        
        # Email body
        body = f"""Dear {customer.get('name', 'Customer')},

Thank you for your order at Gulmohur!

Please find attached your order receipt for Order #{order.get('order_number', 'N/A')}.

Order Summary:
- Total Amount: ₹{order.get('price', 0):.0f}
- Advance Received: ₹{order.get('advance_received', 0):.0f}
- Balance Due: ₹{order.get('balance_amount', 0):.0f}
- Delivery Date: {order.get('delivery_date', 'N/A')}

For payments, you can use UPI: 9711217193@ptsbi (Nitin Gupta)

If you have any questions, please contact us at:
Phone: {SHOP_DETAILS['phone1']}, {SHOP_DETAILS['phone2']}

Thank you for choosing Gulmohur!

Best regards,
{SHOP_DETAILS['owner_name']}
Gulmohur - Fabric for all Seasons
{SHOP_DETAILS['address']}
"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach PDF
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(buffer.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename=receipt_order_{order.get("order_number", order_id)}.pdf')
        msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
        server.send_message(msg)
        server.quit()
        
        return jsonify({'success': True, 'message': 'Receipt sent successfully!'})
        
    except Exception as e:
        print(f"Error sending email: {e}")
        from flask import jsonify
        return jsonify({'success': False, 'message': f'Failed to send email: {str(e)}'}), 500
    
def delete_image_from_firebase(image_url):
    """Delete image from Firebase Storage"""
    if not bucket or not image_url:
        return False
    
    try:
        blob_name = image_url.split(f"{bucket.name}/")[-1]
        blob = bucket.blob(blob_name)
        blob.delete()
        return True
    except Exception as e:
        print(f"Error deleting from Firebase Storage: {e}")
        return False

def fetch_all_orders(customer_id=None):
    """Fetches all orders, optionally filtered by customer_id."""
    orders_ref = get_collection_ref('orders')
    if not orders_ref: return []
    
    query = orders_ref.where('is_deleted', '==', False)

    if customer_id:
        query = query.where('customer_id', '==', customer_id)
        print(f"Showing orders for customer ID: {customer_id}", 'info')
        
    try:
        docs = query.stream()
        orders = []
        for doc in docs:
            order = doc.to_dict()
            order['id'] = doc.id
            orders.append(order)
        return orders
    except Exception as e:
        flash(f"Error fetching orders: {e}", 'error')
        return []

def get_tasks_for_today():
    """Fetches simple placeholder tasks for the dashboard."""
    today_str = datetime.now().strftime('%Y-%m-%d')
    tomorrow_str = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

    tasks = [
        {'title': 'Cut fabric for Maria Sharma Order', 'description': 'Dress Type: Salwar Kameez', 'status': 'due'},
        {'title': 'Final fitting with Ms. Anjali', 'description': '11:00 AM', 'status': 'due'},
        {'title': 'Order thread and buttons', 'description': 'Supplier call required', 'status': 'done'},
        {'title': f'Prepare delivery for Order #2345 (Due: {today_str})', 'description': 'Final checks', 'status': 'due'},
        {'title': f'Follow up on Order #2346 (Due: {tomorrow_str})', 'description': 'Check stitching status', 'status': 'due'},
    ]
    return tasks

from flask import request, flash, redirect, url_for, render_template
from datetime import datetime, timedelta
import json

def clean_for_firestore(obj):
    """Recursively clean data for Firestore compatibility."""
    if isinstance(obj, dict):
        return {k: clean_for_firestore(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        clean_list = []
        for v in obj:
            if isinstance(v, (str, int, float, bool)) or v is None:
                clean_list.append(v)
            elif isinstance(v, dict):
                clean_list.append(json.dumps(v, default=str))
            else:
                clean_list.append(str(v))
        return clean_list
    elif isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj
    else:
        return str(obj)

# -----------------------------
# PROCESS ORDER FORM
# -----------------------------
def process_order_form(order_id=None, customer_id=None, customer_name=''):
    """Create or update an order with images and cost breakdown."""
    form = dict(request.form)
    measurements = parse_measurements(form)

    customer = get_customer_by_id(customer_id) if customer_id else None
    customer_name = customer.get('name') if customer else form.get('customer_name', '')

    # Handle cost breakdown
    material_cost = float(form.get('material_cost') or 0)
    stitching_cost = float(form.get('stitching_cost') or 0)
    total_price = material_cost + stitching_cost
    advance_received = float(form.get('advance_received') or 0)
    balance_amount = total_price - advance_received


    # Get existing images if updating
    existing_images = []
    if order_id:
        existing_order = get_order_by_id(order_id)
        if existing_order:
            existing_images = existing_order.get('images', [])

    # Handle new image uploads
    uploaded_images = []
    if 'order_images' in request.files:
        files = request.files.getlist('order_images')
        for file in files:
            if file and file.filename and allowed_file(file.filename):
                temp_order_id = order_id or f"temp_{datetime.now().timestamp()}"
                image_url = upload_image_to_firebase(file, temp_order_id)
                if image_url:
                    uploaded_images.append(image_url)

    # Combine existing and new images
    all_images = existing_images + uploaded_images

    data = {
        'customer_id': customer_id or form.get('customer_id'),
        'customer_name': customer_name,
        'dress_type': form.get('dress_type'),
        'occasion': form.get('occasion'),
        'fabric': form.get('fabric'),
        'lining': form.get('lining'),
        'delivery_date': form.get('delivery_date'),
        'price': total_price,
        'material_cost': material_cost,
        'stitching_cost': stitching_cost,
        'advance_received': advance_received,
        'balance_amount': balance_amount,
        'notes': form.get('notes'),
        'status': form.get('status', 'Pending'),
        'reference_links': [l for l in request.form.getlist('reference_link') if l.strip()],
        'measurements': measurements,
        'images': all_images,
        'updated_at': firestore.SERVER_TIMESTAMP,
    }

    data = clean_for_firestore(data)
    orders_ref = db.collection('orders')

    if not order_id:
        try:
            docs = list(orders_ref.stream())
            existing_numbers = []
            for d in docs:
                doc_data = d.to_dict()
                num = doc_data.get('order_number')
                if isinstance(num, (int, float)):
                    existing_numbers.append(int(num))
            next_order_number = max(existing_numbers, default=1000) + 1
        except Exception as e:
            print(f"Error generating order number: {e}")
            next_order_number = 1001

        data['order_number'] = next_order_number
        data['created_at'] = firestore.SERVER_TIMESTAMP
        data['is_deleted'] = False

        doc_ref = orders_ref.document()
        doc_ref.set(data)
        order = data.copy()
        order['id'] = doc_ref.id
        flash(f"Order #{next_order_number} created successfully!", "success")
        return order
    else:
        order_doc_ref = orders_ref.document(order_id)
        order_doc = order_doc_ref.get()
        if not order_doc.exists:
            flash(f"No order found with ID {order_id}", "error")
            return None
        order_doc_ref.update(data)
        order = data.copy()
        order['id'] = order_id
        flash("Order updated successfully", "success")
        return order

def parse_measurements(form):
    """Extract measurement fields matching the receipt format"""
    measurements = {
        'length': form.get('length', '').strip(),
        'shoulder': form.get('shoulder', '').strip(),
        'armhole': form.get('armhole', '').strip(),
        'upper_chest': form.get('upper_chest', '').strip(),
        'chest': form.get('chest', '').strip(),
        'waist': form.get('waist', '').strip(),
        'stomach': form.get('stomach', '').strip(),
        'hips': form.get('hips', '').strip(),
        'front_back_cross': form.get('front_back_cross', '').strip(),
        'sleeve_length': form.get('sleeve_length', '').strip(),
        'neck_front_back': form.get('neck_front_back', '').strip(),
        'dart_point': form.get('dart_point', '').strip(),
    }
    
    # Pants measurements
    pants = {
        'length': form.get('pants_length', '').strip(),
        'waist': form.get('pants_waist', '').strip(),
        'thigh': form.get('pants_thigh', '').strip(),
        'mori': form.get('pants_mori', '').strip(),
        'calf': form.get('pants_calf', '').strip(),
    }
    
    measurements['pants'] = pants
    return measurements

# --- Routes ---

@app.route('/')
def dashboard():
    """Main dashboard showing key metrics."""
    orders = fetch_all_orders()
    
    today = datetime.now().date()
    
    pending_orders = 0
    todays_tasks = 0
    total_customers = 0

    if orders:
        pending_orders = sum(1 for o in orders if o.get('status') in ['Pending', 'In Progress'])
        
        def is_due_soon(order):
            try:
                delivery_date = datetime.strptime(order.get('delivery_date', '2999-01-01'), '%Y-%m-%d').date()
                if order.get('status') in ['Pending', 'In Progress']:
                    delta = delivery_date - today
                    return timedelta(days=0) <= delta <= timedelta(days=3)
            except:
                return False
            return False
            
        todays_tasks = sum(1 for o in orders if is_due_soon(o))
    
    customers_ref = get_collection_ref('customers')
    if customers_ref:
        try:
            active_customers = customers_ref.where('is_deleted', '==', False).stream()
            total_customers = len(list(active_customers))
        except Exception as e:
            print(f"Error counting customers: {e}")

    return render_template('dashboard.html', 
        pending_orders=pending_orders, 
        todays_tasks=todays_tasks,
        total_customers=total_customers
    )

# --- Customer Management ---
@app.route('/view_order/<order_id>')
def view_order(order_id):
    """Display all order details including saved measurements."""
    order_ref = db.collection('orders').document(order_id)
    doc = order_ref.get()
    if not doc.exists:
        flash("Order not found", "error")
        return redirect(url_for('all_orders'))

    order = doc.to_dict()
    order['id'] = doc.id
    return render_template('view_order.html', order=order)


@app.route('/new_order/<customer_id>', methods=['GET', 'POST'])
def new_order_for_customer(customer_id):
    """Create a new order for a specific customer with measurements copied from profile."""
    customer_ref = db.collection('customers').document(customer_id)
    customer_doc = customer_ref.get()
    if not customer_doc.exists:
        flash("Customer not found", "error")
        return redirect(url_for('customers'))

    customer = customer_doc.to_dict()
    customer['id'] = customer_doc.id

    if request.method == 'POST':
        return process_order_form(customer_id=customer_id)

    measurements = customer.get('measurements', {})
    order_data = {
        'customer_id': customer_id,
        'customer_name': customer.get('name'),
        'measurements': measurements
    }
    return render_template('order_form.html', title="New Order", customer=customer, order=order_data)

@app.route('/customers')
def customers():
    """List and search all ACTIVE customers (exclude deleted)."""
    customers_ref = get_collection_ref('customers')
    if not customers_ref:
        return render_template('customers.html', customers=[], query='', deleted_count=0)

    query = request.args.get('q', '').strip()
    
    try:
        all_customers = []
        deleted_count = 0
        for doc in customers_ref.stream():
            customer_data = doc.to_dict()
            if customer_data.get('is_deleted', False):
                deleted_count += 1
                continue
            customer_data['id'] = doc.id
            all_customers.append(customer_data)
        
        if query:
            search_results = [
                c for c in all_customers 
                if query.lower() in c.get('name', '').lower() or 
                   query.lower() in c.get('phone', '').lower() or 
                   query.lower() in c.get('email', '').lower()
            ]
            customers_list = search_results
        else:
            customers_list = all_customers

        customers_list.sort(key=lambda c: c.get('name', '').lower())
        
        return render_template('customers.html', customers=customers_list, query=query, deleted_count=deleted_count)
    except Exception as e:
        flash(f"Error retrieving customers: {e}", 'error')
        return render_template('customers.html', customers=[], query=query, deleted_count=0)

@app.route('/customers/deleted')
def deleted_customers():
    """List all deleted customers with option to restore."""
    customers_ref = get_collection_ref('customers')
    if not customers_ref:
        return render_template('deleted_customers.html', customers=[])

    try:
        deleted_customers = []
        for doc in customers_ref.stream():
            customer_data = doc.to_dict()
            if customer_data.get('is_deleted', False):
                customer_data['id'] = doc.id
                deleted_customers.append(customer_data)
        
        deleted_customers.sort(key=lambda c: c.get('deleted_at', ''), reverse=True)
        
        return render_template('deleted_customers.html', customers=deleted_customers)
    except Exception as e:
        flash(f"Error retrieving deleted customers: {e}", 'error')
        return render_template('deleted_customers.html', customers=[])

@app.route('/customer/restore/<customer_id>', methods=['POST'])
def restore_customer(customer_id):
    """Restore a soft-deleted customer."""
    try:
        customers_ref = get_collection_ref('customers')
        customer = get_customer_by_id(customer_id)
        
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for('deleted_customers'))
        
        if not customer.get('is_deleted', False):
            flash("Customer is not deleted.", "warning")
            return redirect(url_for('customers'))
        
        customers_ref.document(customer_id).update({
            'is_deleted': False,
            'restored_at': firestore.SERVER_TIMESTAMP
        })
        
        flash(f"Customer '{customer.get('name', '')}' has been restored successfully!", "success")
        return redirect(url_for('view_customer', customer_id=customer_id))
        
    except Exception as e:
        flash(f"Error restoring customer: {e}", "error")
        return redirect(url_for('deleted_customers'))

import re

@app.route('/customer/add', methods=['GET', 'POST'])
def add_customer():
    """Add a new customer – prevents duplicate phone numbers and assigns numeric ID."""
    customers_ref = get_collection_ref('customers')
    if not customers_ref:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        form = request.form
        name = form.get('name', '').strip()
        phone = form.get('phone', '').strip()
        email = form.get('email', '').strip()

        normalized_phone = re.sub(r'\D', '', phone)

        try:
            existing_customers = customers_ref.stream()
            for doc in existing_customers:
                c = doc.to_dict()
                if c.get('is_deleted', False):
                    continue
                existing_phone = re.sub(r'\D', '', c.get('phone', '') or '')
                if existing_phone == normalized_phone:
                    flash(f"A customer with phone number {phone} already exists ({c.get('name', 'Unnamed')}).", 'error')
                    return redirect(url_for('add_customer'))
        except Exception as e:
            flash(f"Error checking existing customers: {e}", 'error')
            return redirect(url_for('add_customer'))

        try:
            docs = list(customers_ref.stream())
            existing_ids = []
            for d in docs:
                data = d.to_dict()
                cid = data.get('customer_numeric_id')
                if isinstance(cid, (int, float)):
                    existing_ids.append(int(cid))
            next_id = max(existing_ids, default=1000) + 1
        except Exception as e:
            flash(f"Error generating customer ID: {e}", 'error')
            next_id = 1001

        customer_data = {
            'customer_numeric_id': next_id,
            'name': name,
            'phone': phone,
            'email': email,
            'address': {
                'house': form.get('house', '').strip(),
                'locality': form.get('locality', '').strip(),
                'city': form.get('city', '').strip(),
                'state': form.get('state', '').strip(),
                'pin': form.get('pin', '').strip(),
            },
            'measurements': parse_measurements(dict(form)),
            'created_at': firestore.SERVER_TIMESTAMP,
            'updated_at': firestore.SERVER_TIMESTAMP,
            'is_deleted': False,
        }

        try:
            _, doc_ref = customers_ref.add(customer_data)
            flash(f"Customer #{next_id} '{customer_data['name']}' added successfully!", 'success')
            return redirect(url_for('view_customer', customer_id=doc_ref.id))
        except Exception as e:
            flash(f"Error adding customer: {e}", 'error')

    return render_template('customer_form.html', customer=None)

@app.route('/customer/edit/<customer_id>', methods=['GET', 'POST'])
def edit_customer(customer_id):
    """Edit an existing customer."""
    customers_ref = get_collection_ref('customers')
    if not customers_ref:
        return redirect(url_for('dashboard'))
        
    customer = get_customer_by_id(customer_id)
    if not customer:
        flash("Customer not found.", 'error')
        return redirect(url_for('customers'))

    if request.method == 'POST':
        form = request.form
        
        updated_data = {
            'name': form.get('name', '').strip(),
            'phone': form.get('phone', '').strip(),
            'email': form.get('email', '').strip(),
            'address': {
                'house': form.get('house', '').strip(),
                'locality': form.get('locality', '').strip(),
                'city': form.get('city', '').strip(),
                'state': form.get('state', '').strip(),
                'pin': form.get('pin', '').strip(),
            },
            'measurements': parse_measurements(form),
            'updated_at': firestore.SERVER_TIMESTAMP,
        }
        
        try:
            customers_ref.document(customer_id).update(updated_data)
            
            orders_ref = get_collection_ref('orders')
            if orders_ref:
                batch = db.batch()
                orders_query = orders_ref.where('customer_id', '==', customer_id).stream() 
                for order_doc in orders_query:
                    batch.update(order_doc.reference, {'customer_name': updated_data['name']})
                batch.commit()

            flash(f"Customer '{updated_data['name']}' updated successfully!", 'success')
            return redirect(url_for('view_customer', customer_id=customer_id))
        except Exception as e:
            flash(f"Error updating customer: {e}", 'error')
            
        customer = updated_data | {'id': customer_id}

    return render_template('customer_form.html', customer=customer)

@app.route('/customer/<customer_id>')
def view_customer(customer_id):
    """View customer profile, including measurements, linked orders, and financials."""
    customer = get_customer_by_id(customer_id)
    if not customer:
        flash("Customer not found.", 'error')
        return redirect(url_for('customers'))

    orders = fetch_all_orders(customer_id=customer_id)

    total_revenue = 0
    total_outstanding = 0
    for o in orders:
        price = float(o.get('price', 0))
        payments = parse_payments(o.get('payments', []))
        paid_amount = sum(float(p.get('amount', 0)) for p in payments)
        balance_due = max(price - paid_amount, 0)
        o['paid_amount'] = paid_amount
        o['balance_due'] = balance_due
        total_revenue += paid_amount
        total_outstanding += balance_due

    customer_financials = {
        'total_revenue': total_revenue,
        'total_outstanding': total_outstanding
    }

    return render_template(
        'view_customer.html',
        customer=customer,
        orders=orders,
        customer_financials=customer_financials
    )

# --- Order Management ---

@app.route('/orders')
def all_orders():
    """List all orders from all customers."""
    orders = fetch_all_orders()
    return render_template('all_orders.html', orders=orders)

@app.route('/order/new/<customer_id>', methods=['GET', 'POST'])
def new_order(customer_id):
    customer = get_customer_by_id(customer_id)
    if not customer:
        flash("Customer not found.", "error")
        return redirect(url_for('customers'))

    if request.method == 'POST':
        order_id = process_order_form(order_id=None, customer_id=customer_id)
        if order_id:
            return redirect(url_for('view_customer', customer_id=customer_id))
        else:
            return redirect(url_for('new_order', customer_id=customer_id))

    default_date = (datetime.now() + timedelta(weeks=2)).strftime('%Y-%m-%d')
    default_order = {
        'delivery_date': default_date,
        'status': 'Pending',
        'reference_links': [],
        'measurements': customer.get('measurements', {'top': {}, 'bottom': {}}),
        'customer_name': customer.get('name', '')
    }

    return render_template(
        'order_form.html',
        customer=customer,
        order=default_order,
        title="Add New Order"
    )

@app.route('/order/clone/<order_id>')
def clone_order(order_id):
    """Clone an existing order with a new order number."""
    original_order = get_order_by_id(order_id)
    orders_ref = get_collection_ref('orders')
    
    if not original_order or not orders_ref:
        flash("Order not found or Database not ready.", 'error')
        return redirect(url_for('dashboard'))

    customer = get_customer_by_id(original_order['customer_id'])
    if not customer:
        flash("Customer not found.", 'error')
        return redirect(url_for('dashboard'))

    try:
        docs = list(orders_ref.stream())
        existing_numbers = []
        for d in docs:
            doc_data = d.to_dict()
            num = doc_data.get('order_number')
            if isinstance(num, (int, float)):
                existing_numbers.append(int(num))
        next_order_number = max(existing_numbers, default=1000) + 1

        cloned_data = {
            'customer_id': original_order['customer_id'],
            'customer_name': original_order.get('customer_name', ''),
            'dress_type': original_order.get('dress_type', ''),
            'occasion': original_order.get('occasion', ''),
            'fabric': original_order.get('fabric', ''),
            'lining': original_order.get('lining', ''),
            'delivery_date': (datetime.now() + timedelta(weeks=2)).strftime('%Y-%m-%d'),
            'price': original_order.get('price', 0),
            'material_cost': original_order.get('material_cost', 0),
            'stitching_cost': original_order.get('stitching_cost', 0),
            'notes': f"Cloned from Order #{original_order.get('order_number', 'N/A')}",
            'status': 'Pending',
            'reference_links': original_order.get('reference_links', []),
            'measurements': original_order.get('measurements', {}),
            'images': original_order.get('images', []),
            'order_number': next_order_number,
            'created_at': firestore.SERVER_TIMESTAMP,
            'updated_at': firestore.SERVER_TIMESTAMP,
            'is_deleted': False,
        }

        cloned_data = clean_for_firestore(cloned_data)
        doc_ref = orders_ref.document()
        doc_ref.set(cloned_data)
        
        flash(f"Order #{next_order_number} created successfully as a clone!", "success")
        return redirect(url_for('edit_order', order_id=doc_ref.id))
        
    except Exception as e:
        flash(f"Error cloning order: {e}", "error")
        return redirect(url_for('view_customer', customer_id=original_order['customer_id']))

@app.route('/order/tailor-sheet/<order_id>')
def generate_tailor_sheet(order_id):
    """Generate a printable PDF worksheet for the tailor"""
    order = get_order_by_id(order_id)
    if not order:
        flash("Order not found", "error")
        return redirect(url_for('all_orders'))
    
    customer = get_customer_by_id(order['customer_id'])
    if not customer:
        flash("Customer not found", "error")
        return redirect(url_for('all_orders'))
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#ad2a51'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#ad2a51'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    elements.append(Paragraph("GULMOHOUR BOUTIQUE", title_style))
    elements.append(Paragraph("Tailor Work Order", styles['Heading2']))
    elements.append(Spacer(1, 20))
    
    order_info = [
        ['Order Number:', f"#{order.get('order_number', 'N/A')}"],
        ['Date:', datetime.now().strftime('%d %B %Y')],
        ['Delivery Date:', order.get('delivery_date', 'N/A')],
        ['Status:', order.get('status', 'Pending')]
    ]
    
    order_table = Table(order_info, colWidths=[2*inch, 4*inch])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fce8eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(order_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("CUSTOMER DETAILS", heading_style))
    customer_info = [
        ['Name:', customer.get('name', 'N/A')],
        ['Phone:', customer.get('phone', 'N/A')],
        ['Customer ID:', f"#{customer.get('customer_numeric_id', 'N/A')}"]
    ]
    
    customer_table = Table(customer_info, colWidths=[2*inch, 4*inch])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fce8eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("GARMENT SPECIFICATIONS", heading_style))
    garment_info = [
        ['Dress Type:', order.get('dress_type', 'N/A')],
        ['Occasion:', order.get('occasion', 'N/A')],
        ['Fabric:', order.get('fabric', 'N/A')],
        ['Lining:', order.get('lining', 'N/A')],
        ['Material Cost:', f"₹{order.get('material_cost', 0)}"],
        ['Stitching Cost:', f"₹{order.get('stitching_cost', 0)}"],
        ['Total Price:', f"₹{order.get('price', 0)}"]
    ]
    
    garment_table = Table(garment_info, colWidths=[2*inch, 4*inch])
    garment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fce8eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(garment_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("MEASUREMENTS (inches)", heading_style))
    
    measurements = order.get('measurements', {})
    
    if measurements:
        elements.append(Paragraph("Main Measurements:", styles['Heading3']))
        main_data = [['Measurement', 'Value']]
        for key, value in measurements.items():
            if key != 'pants' and value:
                main_data.append([key.replace('_', ' ').title(), f"{value}\""])
        
        if len(main_data) > 1:
            main_table = Table(main_data, colWidths=[3*inch, 3*inch])
            main_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ad2a51')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(main_table)
            elements.append(Spacer(1, 15))
    
    pants = measurements.get('pants', {})
    if pants:
        elements.append(Paragraph("Pants Measurements:", styles['Heading3']))
        pants_data = [['Measurement', 'Value']]
        for key, value in pants.items():
            if value:
                pants_data.append([key.replace('_', ' ').title(), f"{value}\""])
        
        if len(pants_data) > 1:
            pants_table = Table(pants_data, colWidths=[3*inch, 3*inch])
            pants_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ad2a51')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(pants_table)
            elements.append(Spacer(1, 20))
    
    if order.get('notes'):
        elements.append(Paragraph("DESIGN NOTES & SPECIAL INSTRUCTIONS", heading_style))
        notes_para = Paragraph(order.get('notes', ''), styles['Normal'])
        elements.append(notes_para)
        elements.append(Spacer(1, 20))
    
    if order.get('images'):
        elements.append(PageBreak())
        elements.append(Paragraph("REFERENCE IMAGES", heading_style))
        elements.append(Spacer(1, 10))
        
        elements.append(Paragraph(
            "Please refer to these design reference images:",
            styles['Normal']
        ))
        elements.append(Spacer(1, 10))
        
        for idx, img_url in enumerate(order.get('images', []), 1):
            elements.append(Paragraph(f"{idx}. {img_url}", styles['Normal']))
            elements.append(Spacer(1, 5))
    
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("_" * 80, styles['Normal']))
    elements.append(Spacer(1, 10))
    
    footer_data = [
        ['Tailor Signature:', '___________________'],
        ['Date Received:', '___________________'],
        ['Expected Completion:', order.get('delivery_date', '___________________')]
    ]
    
    footer_table = Table(footer_data, colWidths=[2.5*inch, 3.5*inch])
    footer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(footer_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"tailor_sheet_order_{order.get('order_number', order_id)}.pdf",
        mimetype='application/pdf'
    )

@app.route('/order/customer-receipt/<order_id>')
def generate_customer_receipt(order_id):
    """Generate a customer receipt PDF matching the provided format"""
    order = get_order_by_id(order_id)
    if not order:
        flash("Order not found", "error")
        return redirect(url_for('all_orders'))
    
    customer = get_customer_by_id(order['customer_id'])
    if not customer:
        flash("Customer not found", "error")
        return redirect(url_for('all_orders'))
    
    buffer = generate_receipt_pdf(order, customer)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"customer_receipt_order_{order.get('order_number', order_id)}.pdf",
        mimetype='application/pdf'
    )

@app.route('/order/delete-image/<order_id>', methods=['POST'])
def delete_order_image(order_id):
    """Delete a specific image from an order"""
    try:
        from flask import request, jsonify
        
        image_url = request.json.get('image_url')
        if not image_url:
            return jsonify({'success': False, 'message': 'Image URL required'}), 400
        
        order = get_order_by_id(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'Order not found'}), 404
        
        images = order.get('images', [])
        if image_url in images:
            images.remove(image_url)
            
            orders_ref = get_collection_ref('orders')
            orders_ref.document(order_id).update({'images': images})
            
            delete_image_from_firebase(image_url)
            
            return jsonify({'success': True, 'message': 'Image deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Image not found'}), 404
            
    except Exception as e:
        print(f"Error deleting image: {e}")
        from flask import jsonify
        return jsonify({'success': False, 'message': str(e)}), 500
        
@app.route('/order/whatsapp-receipt/<order_id>')
def whatsapp_receipt(order_id):
    """Generate WhatsApp message with receipt details"""
    order = get_order_by_id(order_id)
    if not order:
        flash("Order not found", "error")
        return redirect(url_for('all_orders'))
    
    customer = get_customer_by_id(order['customer_id'])
    if not customer:
        flash("Customer not found", "error")
        return redirect(url_for('all_orders'))
    
    message = f"""*Gulmohur - Order Confirmation*

Dear {customer.get('name', 'Customer')},

Thank you for your order!

*Order Details:*
Order No: #{order.get('order_number', 'N/A')}
Dress Type: {order.get('dress_type', 'N/A')}
Delivery Date: {order.get('delivery_date', 'N/A')}

*Payment Details:*
Total Amount: ₹{order.get('price', 0):.0f}
Advance Received: ₹{order.get('advance_received', 0):.0f}
Balance Due: ₹{order.get('balance_amount', 0):.0f}

*Shop Details:*
{SHOP_DETAILS['shop_name']}
{SHOP_DETAILS['tagline']}
{SHOP_DETAILS['address']}
Phone: {SHOP_DETAILS['phone1']}, {SHOP_DETAILS['phone2']}

For any queries, please contact us.

Thank you!
- Team Gulmohur"""
    
    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    
    customer_phone = customer.get('phone', '').replace(' ', '').replace('-', '').replace('+', '')
    if not customer_phone.startswith('91'):
        customer_phone = '91' + customer_phone
    
    whatsapp_url = f"https://wa.me/{customer_phone}?text={encoded_message}"
    
    from flask import redirect as flask_redirect
    return flask_redirect(whatsapp_url)

@app.route('/order/edit/<order_id>', methods=['GET', 'POST'])
def edit_order(order_id):
    """Edit an existing order - prevent editing for deleted customers."""
    order = get_order_by_id(order_id)
    orders_ref = get_collection_ref('orders')
    
    if not order or not orders_ref:
        flash("Order not found or Database not ready.", 'error')
        return redirect(url_for('dashboard'))

    customer = get_customer_by_id(order['customer_id'])
    
    if customer and customer.get('is_deleted', False):
        flash("Cannot edit orders for deleted customers. Restore the customer first.", 'error')
        return redirect(url_for('view_customer', customer_id=order['customer_id']))
    
    customer_name = customer['name'] if customer else ''

    if request.method == 'POST':
        updated_data = process_order_form(order_id=order_id, customer_id=order['customer_id'], customer_name=customer_name)
        if updated_data:
            return redirect(url_for('view_customer', customer_id=order['customer_id']))
        else:
            flash("Error updating order", "error")
            
    if order.get('delivery_date'):
        try:
            datetime.strptime(order['delivery_date'], '%Y-%m-%d')
        except ValueError:
            order['delivery_date'] = (datetime.now() + timedelta(weeks=2)).strftime('%Y-%m-%d')

    return render_template('order_form.html', customer=customer, order=order, title="Edit Order")

@app.route('/order/delete/<order_id>', methods=['POST'])
def delete_order(order_id):
    """Delete an order - prevent deletion for deleted customers."""
    order = get_order_by_id(order_id)
    orders_ref = get_collection_ref('orders')

    if not order or not orders_ref:
        flash("Order not found or Database not ready.", 'error')
        return redirect(url_for('dashboard'))

    customer_id = order['customer_id']
    customer = get_customer_by_id(customer_id)
    
    if customer and customer.get('is_deleted', False):
        flash("Cannot delete orders for deleted customers. Restore the customer first.", 'error')
        return redirect(url_for('view_customer', customer_id=customer_id))
    
    try:
        orders_ref.document(order_id).update({'is_deleted': True})
        flash(f"Order has been deleted.", 'success')
    except Exception as e:
        flash(f"Error deleting order: {e}", 'error')

    if customer_id:
        return redirect(url_for('view_customer', customer_id=customer_id))
    else:
        return redirect(url_for('all_orders'))
    
@app.route('/tasks')
def daily_tasks():
    """Shows orders due today or in the next 3 days."""
    orders = fetch_all_orders()
    today = datetime.now().date()
    
    def is_due_soon(order):
        if order.get('status') in ['Pending', 'In Progress']:
            try:
                delivery_date = datetime.strptime(order.get('delivery_date', '2999-01-01'), '%Y-%m-%d').date()
                delta = delivery_date - today
                return timedelta(days=0) <= delta <= timedelta(days=3)
            except:
                return False
        return False
        
    due_orders = [o for o in orders if is_due_soon(o)]
    due_orders.sort(key=lambda o: datetime.strptime(o['delivery_date'], '%Y-%m-%d').date())
    
    return render_template('tasks.html', 
        due_orders=due_orders, 
        today=today.strftime('%A, %B %d, %Y')
    )

@app.route('/reports')
def reports_dashboard():
    """Show dashboard of all customers with links to their reports."""
    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    return render_template('reports_dashboard.html', customers=customers)

@app.route('/reports/customer/<customer_id>')
def customer_report(customer_id):
    """Generate financial report for a single customer."""
    customer = get_customer_by_id(customer_id)
    if not customer:
        flash("Customer not found", "error")
        return redirect(url_for('reports_dashboard'))

    orders = fetch_all_orders(customer_id=customer_id)

    total_revenue = sum(float(o.get('price', 0)) for o in orders)
    total_paid = sum(sum(float(p.get('amount', 0)) for p in parse_payments(o.get('payments', []))) for o in orders)
    total_outstanding = total_revenue - total_paid

    graph_data = {
        'orders': [o['id'] for o in orders],
        'revenue': [float(o.get('price', 0)) for o in orders],
        'paid': [sum(float(p.get('amount', 0)) for p in parse_payments(o.get('payments', []))) for o in orders],
        'balance': [float(o.get('price', 0)) - sum(float(p.get('amount', 0)) for p in parse_payments(o.get('payments', []))) for o in orders]
    }

    return render_template('report_customer.html',
                           customer=customer,
                           total_revenue=total_revenue,
                           total_paid=total_paid,
                           total_outstanding=total_outstanding,
                           graph_data=graph_data)

@app.route('/reports/boutique')
def boutique_report():
    """Generate overall boutique financial report."""
    orders = fetch_all_orders()
    customers_ref = get_collection_ref('customers')

    total_revenue = sum(float(o.get('price', 0)) for o in orders)
    total_paid = sum(sum(float(p.get('amount', 0)) for p in parse_payments(o.get('payments', []))) for o in orders)
    total_outstanding = total_revenue - total_paid

    graph_data = {'customers': [], 'revenue': [], 'paid': [], 'balance': []}
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if c.get('is_deleted', False):
                continue
            c_id = doc.id
            c_orders = fetch_all_orders(customer_id=c_id)
            c_revenue = sum(float(o.get('price', 0)) for o in c_orders)
            c_paid = sum(sum(float(p.get('amount', 0)) for p in parse_payments(o.get('payments', []))) for o in c_orders)
            c_balance = c_revenue - c_paid
            graph_data['customers'].append(c.get('name', 'Unknown'))
            graph_data['revenue'].append(c_revenue)
            graph_data['paid'].append(c_paid)
            graph_data['balance'].append(c_balance)

    return render_template('report_boutique.html',
                           total_revenue=total_revenue,
                           total_paid=total_paid,
                           total_outstanding=total_outstanding,
                           graph_data=graph_data)

@app.route('/customer/delete/<customer_id>', methods=['POST'])
def delete_customer(customer_id):
    """Smart delete logic for customers."""
    try:
        customers_ref = get_collection_ref('customers')
        orders_ref = get_collection_ref('orders')

        customer = get_customer_by_id(customer_id)
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for('customers'))

        orders = []
        if orders_ref:
            query = orders_ref.where('customer_id', '==', customer_id).where('is_deleted', '==', False)
            orders = [doc.to_dict() | {'id': doc.id} for doc in query.stream()]

        open_orders = [o for o in orders if o.get('status') in ['Pending', 'In Progress']]

        if open_orders:
            flash(f"Cannot delete customer '{customer.get('name', '')}' because they have open orders.", "error")
            return redirect(url_for('view_customer', customer_id=customer_id))
        else:
            customers_ref.document(customer_id).update({
                'is_deleted': True,
                'deleted_at': firestore.SERVER_TIMESTAMP
            })
            flash(f"Customer '{customer.get('name', '')}' has been deleted and can be restored from the Deleted Customers page.", "warning")
            return redirect(url_for('customers'))

    except Exception as e:
        flash(f"Error deleting customer: {e}", "error")

    return redirect(url_for('customers'))

# ============================================================
# CUSTOMER LIST REPORT
# ============================================================

@app.route('/reports/customer-list')
def customer_list_report():
    """Show customer list report page with phone numbers."""
    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    customers.sort(key=lambda c: c.get('name', '').lower())
    return render_template('customer_list_report.html', customers=customers)


@app.route('/reports/customer-list/pdf')
def customer_list_report_pdf():
    """Download customer list as PDF."""
    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    customers.sort(key=lambda c: c.get('name', '').lower())

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Normal'],
                                  fontSize=18, alignment=TA_CENTER,
                                  textColor=colors.HexColor('#ad2a51'),
                                  spaceAfter=4, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
                                fontSize=9, alignment=TA_CENTER,
                                textColor=colors.grey, spaceAfter=16)

    elements.append(Paragraph(SHOP_DETAILS['shop_name'], title_style))
    elements.append(Paragraph('Customer List Report', sub_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y')}", sub_style))
    elements.append(Spacer(1, 10))

    # Table header + rows
    table_data = [['#', 'Customer ID', 'Name', 'Phone', 'Email']]
    for idx, c in enumerate(customers, 1):
        table_data.append([
            str(idx),
            str(c.get('customer_numeric_id', 'N/A')),
            c.get('name', 'N/A'),
            c.get('phone', 'N/A'),
            c.get('email', 'N/A') or '-',
        ])

    col_widths = [0.4*inch, 0.9*inch, 2.2*inch, 1.4*inch, 2.3*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ad2a51')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fce8eb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"Total Customers: {len(customers)}", 
                               ParagraphStyle('Footer', parent=styles['Normal'],
                                              fontSize=9, fontName='Helvetica-Bold')))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"customer_list_{datetime.now().strftime('%Y%m%d')}.pdf",
                     mimetype='application/pdf')


@app.route('/reports/customer-list/xls')
def customer_list_report_xls():
    """Download customer list as XLS (Excel)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    customers.sort(key=lambda c: c.get('name', '').lower())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer List"

    # Title row
    ws.merge_cells('A1:E1')
    ws['A1'] = f"{SHOP_DETAILS['shop_name']} - Customer List Report"
    ws['A1'].font = Font(bold=True, size=14, color='AD2A51')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%d %B %Y')}"
    ws['A2'].font = Font(size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Header row
    headers = ['#', 'Customer ID', 'Name', 'Phone', 'Email']
    header_fill = PatternFill('solid', start_color='AD2A51')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    alt_fill = PatternFill('solid', start_color='FCE8EB')
    for idx, c in enumerate(customers, 1):
        row = idx + 4
        row_fill = alt_fill if idx % 2 == 0 else None
        row_data = [
            idx,
            c.get('customer_numeric_id', 'N/A'),
            c.get('name', 'N/A'),
            c.get('phone', 'N/A'),
            c.get('email', '') or '',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(size=10)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if row_fill:
                cell.fill = row_fill

    # Total row
    total_row = len(customers) + 5
    ws.cell(row=total_row, column=1, value='Total')
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=2, value=len(customers))
    ws.cell(row=total_row, column=2).font = Font(bold=True, size=10)

    # Column widths
    col_widths_map = {'A': 6, 'B': 13, 'C': 28, 'D': 18, 'E': 32}
    for col_letter, width in col_widths_map.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[4].height = 22

    import tempfile, os

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)
    
    @after_this_request
    def remove_file(response):
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        return response
    
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f"customer_list_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/reports/customer-list/whatsapp')
def customer_list_report_whatsapp():
    """Send customer list summary via WhatsApp to the shop owner."""
    import urllib.parse

    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    customers.sort(key=lambda c: c.get('name', '').lower())

    lines = [f"*{SHOP_DETAILS['shop_name']} - Customer List*",
             f"_{datetime.now().strftime('%d %B %Y')}_",
             f"Total Customers: {len(customers)}", ""]

    for idx, c in enumerate(customers, 1):
        lines.append(f"{idx}. {c.get('name', 'N/A')} - {c.get('phone', 'N/A')}")

    message = '\n'.join(lines)
    encoded = urllib.parse.quote(message)
    business_number = WHATSAPP_CONFIG['business_number']
    return redirect(f"https://wa.me/{business_number}?text={encoded}")


@app.route('/reports/customer-list/email', methods=['POST'])
def customer_list_report_email():
    """Email customer list PDF to the provided email address."""
    from flask import jsonify
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    recipient_email = request.json.get('email', '').strip()
    if not recipient_email:
        return jsonify({'success': False, 'message': 'Email address required'}), 400

    customers_ref = get_collection_ref('customers')
    customers = []
    if customers_ref:
        for doc in customers_ref.stream():
            c = doc.to_dict()
            if not c.get('is_deleted', False):
                c['id'] = doc.id
                customers.append(c)
    customers.sort(key=lambda c: c.get('name', '').lower())

    # Generate PDF inline (reuse PDF logic)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Normal'],
                                  fontSize=18, alignment=TA_CENTER,
                                  textColor=colors.HexColor('#ad2a51'),
                                  spaceAfter=4, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
                                fontSize=9, alignment=TA_CENTER,
                                textColor=colors.grey, spaceAfter=16)

    elements.append(Paragraph(SHOP_DETAILS['shop_name'], title_style))
    elements.append(Paragraph('Customer List Report', sub_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y')}", sub_style))
    elements.append(Spacer(1, 10))

    table_data = [['#', 'Customer ID', 'Name', 'Phone', 'Email']]
    for idx, c in enumerate(customers, 1):
        table_data.append([str(idx), str(c.get('customer_numeric_id', 'N/A')),
                           c.get('name', 'N/A'), c.get('phone', 'N/A'),
                           c.get('email', '') or '-'])

    col_widths = [0.4*inch, 0.9*inch, 2.2*inch, 1.4*inch, 2.3*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ad2a51')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fce8eb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"Total Customers: {len(customers)}",
                               ParagraphStyle('Footer', parent=styles['Normal'],
                                              fontSize=9, fontName='Helvetica-Bold')))
    doc.build(elements)
    buffer.seek(0)

    try:
        msg = MIMEMultipart()
        msg['From'] = f"{EMAIL_CONFIG['sender_name']} <{EMAIL_CONFIG['sender_email']}>"
        msg['To'] = recipient_email
        msg['Subject'] = f"Customer List - {SHOP_DETAILS['shop_name']} - {datetime.now().strftime('%d %b %Y')}"

        body = f"""Please find attached the customer list for {SHOP_DETAILS['shop_name']}.

Total Customers: {len(customers)}
Generated on: {datetime.now().strftime('%d %B %Y')}

Best regards,
{SHOP_DETAILS['owner_name']}
{SHOP_DETAILS['shop_name']}"""
        msg.attach(MIMEText(body, 'plain'))

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(buffer.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        f'attachment; filename=customer_list_{datetime.now().strftime("%Y%m%d")}.pdf')
        msg.attach(part)

        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
        server.send_message(msg)
        server.quit()

        return jsonify({'success': True, 'message': f'Customer list sent to {recipient_email}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to send email: {str(e)}'}), 500
        
from flask import jsonify

@app.route('/api/ai-suggestions', methods=['POST'])
def ai_suggestions():
    """Generate AI design suggestions based on order details"""
    try:
        data = request.json
        dress_type = data.get('dress_type', '')
        occasion = data.get('occasion', '')
        fabric = data.get('fabric', '')
        
        suggestions = generate_design_suggestions(dress_type, occasion, fabric)
        
        return jsonify({'suggestions': suggestions})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_design_suggestions(dress_type, occasion, fabric):
    """Generate design suggestions based on inputs"""
    suggestions = []
    
    if 'lehenga' in dress_type.lower():
        suggestions.append("Consider heavy embroidery on the border")
        suggestions.append("Add mirror work or zari for traditional look")
    elif 'saree' in dress_type.lower():
        suggestions.append("Designer blouse with contrasting colors works well")
        suggestions.append("Consider embroidered borders")
    elif 'salwar' in dress_type.lower():
        suggestions.append("Straight cut suits are trending")
        suggestions.append("Add gota patti work for festive occasions")
    elif 'kurti' in dress_type.lower():
        suggestions.append("A-line kurtis are flattering")
        suggestions.append("Consider block prints or digital prints")
    
    if 'wedding' in occasion.lower():
        suggestions.append("Use rich colors like red, maroon, or royal blue")
        suggestions.append("Heavy embellishments and stone work recommended")
    elif 'party' in occasion.lower():
        suggestions.append("Contemporary designs with sequins work well")
        suggestions.append("Consider trendy colors like pastels or metallics")
    elif 'casual' in occasion.lower():
        suggestions.append("Lightweight fabrics with minimal work")
        suggestions.append("Comfortable fit with simple designs")
    
    if 'silk' in fabric.lower():
        suggestions.append("Silk fabric pairs well with gold zari work")
        suggestions.append("Keep design elegant to showcase fabric quality")
    elif 'cotton' in fabric.lower():
        suggestions.append("Light embroidery suits cotton fabric")
        suggestions.append("Consider block printing or hand painting")
    elif 'georgette' in fabric.lower():
        suggestions.append("Georgette works well with flowing designs")
        suggestions.append("Add sequin or stone work for shimmer")
    
    if not suggestions:
        suggestions.append(f"For {dress_type}, consider traditional embroidery patterns")
        suggestions.append("Add embellishments according to the occasion")
    
    return " • ".join(suggestions)

@app.route('/orders-by-period')
def orders_by_period():
    """View orders filtered by date period - only show results after search"""
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    
    # ✅ Only fetch orders if dates are provided
    if not from_date or not to_date:
        return render_template('orders_by_period.html', 
                             orders=None,  # None means "no search yet"
                             from_date='', 
                             to_date='')
    
    orders_ref = get_collection_ref('orders')
    if not orders_ref:
        return render_template('orders_by_period.html', orders=[], from_date=from_date, to_date=to_date)
    
    query = orders_ref.where('is_deleted', '==', False)
    
    try:
        all_orders = []
        for doc in query.stream():
            order = doc.to_dict()
            order['id'] = doc.id
            
            # Calculate balance_amount if not present
            if 'balance_amount' not in order:
                price = float(order.get('price', 0))
                advance = float(order.get('advance_received', 0))
                order['balance_amount'] = price - advance
            
            all_orders.append(order)
        
        # Filter by date
        filtered_orders = []
        from_datetime = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_datetime = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        for order in all_orders:
            delivery_date_str = order.get('delivery_date', '')
            if delivery_date_str:
                try:
                    delivery_date = datetime.strptime(delivery_date_str, '%Y-%m-%d').date()
                    if from_datetime <= delivery_date <= to_datetime:
                        filtered_orders.append(order)
                except:
                    continue
        
        # Sort by delivery date
        filtered_orders.sort(key=lambda x: x.get('delivery_date', '9999-12-31'))
        
        return render_template('orders_by_period.html', 
                             orders=filtered_orders, 
                             from_date=from_date, 
                             to_date=to_date)
    except Exception as e:
        flash(f"Error fetching orders: {e}", 'error')
        return render_template('orders_by_period.html', orders=[], from_date=from_date, to_date=to_date)

if __name__ == "__main__":
    app.run()
